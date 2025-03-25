import pandas as pd
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.drawing.image import Image  # Linha adicionada
import matplotlib.pyplot as plt
from openpyxl.chart import BarChart, Reference
import xlsxwriter

# Lista para armazenar os dados dos enxadristas
enxadristas = []
RATING_INICIAL = 1600
K = 32  # Fator de ajuste do rating Elo


# Funções para calcular expectativa de vitória e atualizar rating
def atualizar_rating(rating_atual, resultado, expectativa):
    return round(rating_atual + K * (resultado - expectativa))

def expectativa_vitoria(rating_jogador, rating_oponente):
    """Calcula a expectativa de vitória do jogador baseado na fórmula de Elo."""
    return 1 / (1 + 10 ** ((rating_oponente - rating_jogador) / 400))

def obter_inteiro_valido(mensagem):
    while True:
        try:
            valor = int(input(mensagem))
            if valor < 0:
                raise ValueError
            return valor
        except ValueError:
            print("Erro! Digite um número inteiro válido.")

def cadastrar_enxadrista():
    enxadrista = {}
    enxadrista['nome'] = input('Nome do enxadrista: ').strip()
    enxadrista['rating'] = RATING_INICIAL
    enxadrista['partidas'] = []
    enxadrista['último_torneio'] = input('Digite a data do último torneio (dd/mmm, ex: 17/fev): ').strip()
    
    medalhas = {medalha: obter_inteiro_valido(f'Quantas medalhas de {medalha} {enxadrista["nome"]} possui? ') for medalha in ['ouro', 'prata', 'bronze']}
    enxadrista.update(medalhas)

    tot_partidas = obter_inteiro_valido(f'Quantas partidas {enxadrista["nome"]} jogou?: ')
    for c in range(tot_partidas):
        pontos = obter_inteiro_valido(f'Quantos pontos na partida {c + 1}? ')
        enxadrista['partidas'].append(pontos)

    return enxadrista

# Loop principal para cadastrar múltiplos enxadristas
while True:
    enxadrista = cadastrar_enxadrista()
    enxadristas.append(enxadrista)

    while True:
        resp = input('Deseja cadastrar outro enxadrista? [S/N] ').strip().upper()
        if resp in 'SN':
            break
        print('Erro! Responda apenas S ou N.')

    if resp == 'N':
        break

# Sistema de confrontos
while True:
    resp = input('Deseja registrar um confronto? [S/N] ').strip().upper()
    if resp == 'N':
        break
    if resp == 'S':
        nome1 = input('Nome do primeiro enxadrista: ').strip()
        nome2 = input('Nome do segundo enxadrista: ').strip()
        jogador1 = next((e for e in enxadristas if e['nome'] == nome1), None)
        jogador2 = next((e for e in enxadristas if e['nome'] == nome2), None)

        if not jogador1 or not jogador2:
            print("Erro! Um dos jogadores não foi encontrado.")
            continue

        resultado = input(f'Resultado da partida ({nome1} venceu = 1, {nome2} venceu = 0, empate = 0.5): ').strip()
        try:
            resultado = float(resultado)
            if resultado not in [0, 0.5, 1]:
                raise ValueError
        except ValueError:
            print("Erro! Insira um valor válido para o resultado.")
            continue

        # Atualização do rating Elo
        expectativa1 = expectativa_vitoria(jogador1['rating'], jogador2['rating'])
        expectativa2 = expectativa_vitoria(jogador2['rating'], jogador1['rating'])
        jogador1['rating'] = atualizar_rating(jogador1['rating'], resultado, expectativa1)
        jogador2['rating'] = atualizar_rating(jogador2['rating'], 1 - resultado, expectativa2)
        print(f"Novo rating de {jogador1['nome']}: {jogador1['rating']}")
        print(f"Novo rating de {jogador2['nome']}: {jogador2['rating']}")

# Cria o DataFrame a partir da lista de enxadristas
df = pd.DataFrame(enxadristas)
df.index.name = "cod"  # Define o nome do índice

df['pontos_por_medalha'] = (df['ouro'] * 3) + (df['prata'] * 2) + (df['bronze'] * 1)
df['media_pontos_por_partida'] = df['partidas'].apply(sum) / df['partidas'].apply(len)

df = df.sort_values(by=['pontos_por_medalha'], ascending=[False])
df.insert(0, 'posição', range(1, len(df) + 1))

# Exibe o DataFrame formatado
print(tabulate(df[['posição', 'nome', 'rating', 'ouro', 'prata', 'bronze', 'pontos_por_medalha', 'media_pontos_por_partida', 'último_torneio']], headers='keys', tablefmt='grid', showindex=True))

df.to_excel("ranking_enxadristas.xlsx", index=True)
print("Tabela salva com sucesso no arquivo 'ranking_enxadristas.xlsx'!")

wb = load_workbook("ranking_enxadristas.xlsx")
sheet = wb.active

workbook = xlsxwriter.Workbook('seu_arquivo.xlsx')
worksheet = workbook.add_worksheet()

import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

# Estrutura do gráfico 
plt.figure(figsize=(3, 3))  # Diminuindo o tamanho do gráfico para 5x5

labels = df['nome']

# Tamanho do gráfico de barras para rating e pontos por medalha
sizes_rating = df['rating']
sizes_medals = df['pontos_por_medalha']

# Cria um gráfico de barras para rating
plt.bar(labels, sizes_rating, color='blue', alpha=0.7)  # Ajustando a cor e a transparência
plt.xticks(rotation=45, fontsize=6)  # Rotaciona os rótulos do eixo x para melhor leitura
plt.ylabel('Rating', fontsize=7)  # Adiciona rótulo ao eixo y
plt.title('CHESS MASTER RATING', fontsize=9)  # Ajustando o tamanho do título para melhor leitura

# Adiciona a legenda com os pontos por medalha
medal_labels = [f"{nome}: {pontos} pontos" for nome, pontos in zip(df['nome'], sizes_medals)]
plt.legend(medal_labels, title="Pontos por Medalha", loc="upper left", bbox_to_anchor=(1, 1), fontsize=6, title_fontsize='8')  # Ajustando a posição e a fonte da legenda

# Salva o gráfico
plt.savefig('grafico_barras.png', bbox_inches='tight')  # Mudando o nome do arquivo para gráfico de barras
plt.close()

labels = df['nome']  # Nomes dos enxadristas
sizes = df['pontos_por_medalha'] + df['rating']

# Adicionando o gráfico de barras ao arquivo Excel
img = Image('grafico_barras.png')  # Mudando o nome do arquivo para gráfico de barras
sheet.add_image(img, 'N15')  # Ajuste a posição conforme necessário

try:
    wb.save("ranking_enxadristas_com_grafico.xlsx")
    print("Gráfico atualizado e arquivo salvo como 'ranking_enxadristas_com_grafico.xlsx'!")
except Exception as e:
    print(f"Ocorreu um erro ao salvar o arquivo: {e}")

print('<< Volte sempre >>')